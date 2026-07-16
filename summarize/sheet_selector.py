"""选择汇总工作表 — 纯业务逻辑（无 GUI）。"""
from __future__ import annotations

from io import BytesIO
import functools
import sys
from pathlib import Path
from typing import List, Optional


def _index_sort_key(p: Path) -> int:
    """从文件名提取序号排序，例如 '1-文件名.xlsx' -> 1。"""
    stem = p.stem
    i = 0
    while i < len(stem) and stem[i].isdigit():
        i += 1
    return int(stem[:i]) if i > 0 else 9999


def find_sample_xlsx(
    download_dir: Path,
    *,
    prefer_name_contains: Optional[str] = None,
    output_prefix: str = "录音师薪资结算结果",
) -> Optional[Path]:
    """在下载目录中找第一个「非汇总结果」的 xlsx 作为样例。"""
    if not download_dir.is_dir():
        return None

    files = sorted(
        (f for f in download_dir.glob("*.xlsx") if not f.stem.startswith(output_prefix)),
        key=_index_sort_key,
    )
    if not files:
        return None

    keyword = (prefer_name_contains or "").strip()
    if keyword:
        for f in files:
            if keyword in f.name:
                return f

    return min(files, key=_index_sort_key)


def list_sheet_names(file_path: str | Path) -> List[str]:
    """以样例 xlsx 读取 sheet 名列表。

    直接解析 xlsx（zip 包）中的 xl/workbook.xml 来获取 sheet 名称，
    完全绕过 openpyxl 样式解析，避免样式异常导致返回空列表。
    """
    import zipfile
    import xml.etree.ElementTree as ET

    try:
        with zipfile.ZipFile(str(file_path), "r") as z:
            with z.open("xl/workbook.xml") as f:
                tree = ET.parse(f)
        root = tree.getroot()
        # 处理命名空间
        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets = root.findall(".//s:sheet", ns)
        return [s.attrib.get("name", "") for s in sheets if s.attrib.get("name")]
    except Exception:
        pass

    try:
        # 回退：使用 openpyxl 只读模式
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        pass

    try:
        # 回退：修复样式后再读取
        from vendor.app import fix_excel_file_style

        with open(file_path, "rb") as f:
            content = BytesIO(f.read())
        fixed = fix_excel_file_style(content)
        from openpyxl import load_workbook as lw

        wb = lw(fixed, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def _parse_sheet_choice(choice: str, names: List[str]) -> Optional[str]:
    if not choice or choice == "0":
        selected = names[0]
        print(f"-> 使用默认工作表: {selected}")
        return selected

    if choice.isdigit():
        idx = int(choice)
        if 1 <= idx <= len(names):
            selected = names[idx - 1]
            print(f"-> 已选择: {selected}")
            return selected
        print(f"[!] 序号无效，有效范围 1-{len(names)}")
        return None

    if choice in names:
        print(f"-> 已选择: {choice}")
        return choice

    print(f"-> 使用自定义工作表名: {choice}")
    return choice


def prompt_sheet_name_console(
    sample_file: Optional[Path],
) -> Optional[str]:
    """
    控制台交互式选择工作表。
    - 输入序号 1-N
    - 或直接输入完整 sheet 名
    - 回车 / 0：使用第一个工作表
    """
    names: List[str] = []
    if sample_file and sample_file.is_file():
        names = list_sheet_names(sample_file)
        print("\n" + "=" * 40)
        print(f"样例文件: {sample_file.name}")
    else:
        print("\n" + "=" * 40)
        print("[!] 下载目录中未找到 xlsx，请直接输入工作表名称")

    if names:
        print("可选工作表 (sheet):")
        for i, name in enumerate(names, 1):
            print(f"  {i}. {name}")
        print("  0 / 回车 = 使用第 1 个工作表")
        print("  也可直接输入工作表完整名称（区分大小写）")
    else:
        print("请直接输入要汇总的工作表名称（区分大小写）")
    print("=" * 40)

    try:
        choice = input("\n请输入序号或工作表名称: ").strip()
    except (EOFError, KeyboardInterrupt):
        print("\n已取消选择工作表")
        return None

    if not names:
        if not choice:
            print("[!] 未输入工作表名称")
            return None
        print(f"-> 使用工作表: {choice}")
        return choice

    return _parse_sheet_choice(choice, names)


def resolve_sheet_name(
    sample_file: Optional[Path],
    *,
    config_sheet: Optional[str] = None,
    cli_sheet: Optional[str] = None,
    prompt_interactive: bool = True,
) -> Optional[str]:
    """
    确定汇总使用的 sheet 名。

    优先级：命令行 --sheet > 配置文件 sheet_name > 交互输入 > None(首个工作表)
    """
    if cli_sheet is not None and cli_sheet.strip():
        name = cli_sheet.strip()
        print(f"工作表(命令行): {name}")
        return name

    if config_sheet and config_sheet.strip():
        name = config_sheet.strip()
        print(f"工作表(配置文件): {name}")
        return name

    if prompt_interactive:
        return prompt_sheet_name_console(sample_file)

    print("工作表: 默认（Excel 第一个工作表）")
    return None
